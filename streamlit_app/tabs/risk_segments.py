"""마케팅·CRM 담당자를 위한 위험고객 세분화 운영 화면.

[Doo 작업]
- 적용 Threshold 기반 CRM 핵심 지표와 고객 우선순위 표시
- 관리 범위·고객 유형 복합 필터와 CustomerID 검색 구현
- 선택 고객의 거래 상태·위험 요인·추천 캠페인 상세 정보 연결
- 현재 필터 결과와 전체 캠페인 대상을 구분한 Excel·CSV 다운로드 구현

기존 XGBoost 모델, 전처리, 이탈 확률 및 고객 유형 계산 로직은 변경하지 않습니다.
"""

from datetime import datetime
from io import BytesIO

import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from customer_scoring import load_customer_table, score_customers, segment
from config import DEFAULT_THRESHOLD


BADGE_BY_TYPE = {
    "이탈 위험 높음": "리텐션 우선",
    "첫 구매 고객": "웰컴 캠페인",
    "장기 구매 주기": "리마인드 예정",
    "정상": "모니터링",
}

CAMPAIGN_BY_TYPE = {
    "이탈 위험 높음": "개인화 리텐션 쿠폰",
    "첫 구매 고객": "웰컴 캠페인",
    "장기 구매 주기": "구매 주기 리마인드",
    "정상": "일반 리텐션 캠페인",
}

PRIORITY_COLORS = {
    "관리": ("#EAF3FF", "#1F6FCC", "🔵"),
    "관찰": ("#F2F4F7", "#667085", "⚪"),
}


def _priority_label(probability: float, threshold: float) -> str:
    """[Doo 작업] 적용 기준과 확률을 함께 사용해 CRM 우선순위를 정합니다."""
    if probability < threshold:
        return "관찰"
    return "관리"


def _customer_type_group(row: pd.Series) -> str:
    """[Doo 작업] 기존 고객 유형을 바꾸지 않고 UI 필터용 범주만 추가합니다."""
    if row["고객유형"] == "첫 구매 고객":
        return "첫 구매 고객"
    if row["고객유형"] == "장기 구매 주기":
        return "장기 구매 주기 고객"
    if row["frequency"] > 1:
        return "반복 구매 고객"
    return "기타 관리 고객"


def _campaign_plan(customer: pd.Series) -> dict:
    """[Doo 작업] 기존 고객 유형 규칙을 실무형 추천 문구로 확장합니다."""
    probability = customer["이탈확률"]
    customer_type = customer["고객유형"]

    if customer_type == "이탈 위험 높음":
        return {
            "name": "개인화 리텐션 쿠폰",
            "target": "2회 이상 구매했고, 평균 구매 간격은 90일 미만이지만 마지막 구매 후 90일 이상 지난 고객",
            "reason": f"평소 구매 주기를 초과했고 현재 이탈 확률이 {probability:.1%}입니다.",
            "actions": [
                "7일 이내 사용 가능한 개인화 쿠폰 발송",
                "이전 구매 상품과 연관된 상품 추천",
                "이메일 또는 앱 푸시로 우선 안내",
            ],
            "effect": "빠른 재방문과 재구매를 유도해 단기 이탈을 방지합니다.",
            "caution": "고객별 구매 이력과 마진을 확인해 쿠폰 할인율과 발송 빈도를 조정합니다.",
        }
    if customer_type == "첫 구매 고객":
        return {
            "name": "웰컴 캠페인",
            "target": "현재까지 구매 횟수가 1회인 고객",
            "reason": f"첫 구매 이후 추가 구매가 확인되지 않았고 현재 이탈 확률이 {probability:.1%}입니다.",
            "actions": [
                "첫 구매 고객 전용 재구매 쿠폰 제공",
                "첫 구매 상품과 연관된 상품 추천",
                "한 달 이내 리텐션 메시지 예약",
            ],
            "effect": "두 번째 구매 장벽을 낮춰 신규 고객을 반복 구매 고객으로 전환합니다.",
            "caution": "첫 구매 직후에는 과도한 메시지를 피하고 상품 배송·사용 시점을 고려합니다.",
        }
    if customer_type == "장기 구매 주기":
        return {
            "name": "구매 주기 리마인드",
            "target": "2회 이상 구매했고 평균 구매 간격이 90일 이상인 장기 주기 고객",
            "reason": "평균 구매 간격이 긴 고객이므로 기존 구매 주기에 맞춘 접근이 적합합니다.",
            "actions": [
                "고객의 과거 구매 주기를 참고해 적절한 시점에 상품 재구매 안내 발송",
                "과도한 조기 할인 발송은 제외",
                "관심 상품 재입고 또는 추천 상품 안내",
            ],
            "effect": "고객 고유의 구매 시점에 맞춰 자연스러운 재구매를 유도합니다.",
            "caution": "구매 주기가 긴 고객을 단순 휴면으로 판단해 너무 이른 할인 메시지를 보내지 않습니다.",
        }
    return {
        "name": "일반 리텐션 캠페인",
        "target": "위 세 유형에 해당하지 않는 2회 이상 구매 고객",
        "reason": f"거래 패턴과 현재 이탈 확률 {probability:.1%}를 함께 고려한 관리가 필요합니다.",
        "actions": [
            "일반 리텐션 메시지 발송",
            "최근 구매 상품 기반 추천 제공",
            "캠페인 반응 여부를 확인해 후속 조치 결정",
        ],
        "effect": "관계를 유지하면서 관심 상품의 추가 구매 가능성을 높입니다.",
        "caution": "일괄 할인보다 최근 구매 상품과 캠페인 반응을 활용해 개인화 수준을 높입니다.",
    }


def _render_campaign_guide():
    """고객 유형별 추천 캠페인의 대상과 실행 목적을 한눈에 설명합니다."""
    with st.expander("추천 캠페인 4종 자세히 보기"):
        st.caption(
            "추천 캠페인은 모델이 새로 분류한 결과가 아니라 고객의 구매 횟수와 구매 주기를 이용한 "
            "룰 기반 실행 제안입니다. 이탈 확률 38% 이상 여부는 캠페인 대상 선정에 별도로 사용합니다."
        )
        guide_rows = [
            ("웰컴 캠페인", "구매 1회", "두 번째 구매 쿠폰·연관 상품 안내", "첫 재구매 전환"),
            ("개인화 리텐션 쿠폰", "평균 구매 간격 90일 미만, 최근 구매 후 90일 이상", "개인화 쿠폰·우선 메시지", "단기 이탈 방지"),
            ("구매 주기 리마인드", "평균 구매 간격 90일 이상", "과거 구매 주기에 맞춘 상품 안내", "자연스러운 재구매 유도"),
            ("일반 리텐션 캠페인", "나머지 반복 구매 고객", "최근 구매 기반 상품 추천", "관계 유지·추가 구매"),
        ]
        st.dataframe(
            pd.DataFrame(guide_rows, columns=["캠페인", "추천 대상", "주요 실행", "기대 효과"]),
            hide_index=True,
            width="stretch",
        )
        st.info(
            "분류 우선순위: 구매 1회 고객을 먼저 '첫 구매 고객'으로 구분한 뒤, "
            "반복 구매 고객을 '이탈 위험 높음 → 장기 구매 주기 → 정상' 순서로 구분합니다. "
            "따라서 한 고객에게 여러 캠페인이 동시에 지정되지 않습니다."
        )
        st.caption(
            "여기서 90일은 고객의 정확한 다음 구매일을 예측한 값이 아니라, 프로젝트의 Target 관찰 기간과 "
            "일치시킨 룰 기반 분류 기준입니다. 실제 발송 시점은 고객별 구매 이력과 함께 검토해야 합니다."
        )


def _risk_factors(customer: pd.Series) -> list[str]:
    """[Doo 작업] SHAP 대신 실제 거래 데이터에 해당하는 위험 요인만 설명합니다."""
    factors = []
    recency = customer.get("recency_days")
    cycle_ratio = customer.get("평소_주기_대비")
    frequency = customer.get("frequency")
    recent_ratio = customer.get("recent_activity_ratio")
    has_return = customer.get("has_return")

    if pd.notna(recency) and recency >= 60:
        factors.append(f"마지막 구매 후 {int(recency)}일이 경과했습니다.")
    if pd.notna(cycle_ratio) and cycle_ratio >= 1.2:
        factors.append(f"평균 구매 간격보다 {cycle_ratio:.1f}배 오래 구매하지 않았습니다.")
    if pd.notna(frequency) and int(frequency) == 1:
        factors.append("구매 횟수가 1회인 첫 구매 고객입니다.")
    if pd.notna(recent_ratio) and recent_ratio == 0:
        factors.append("최근 90일 동안 재구매 활동이 확인되지 않았습니다.")
    if has_return == 1:
        factors.append("취소 또는 반품 경험이 확인됩니다.")

    return factors


def _customer_interpretation(customer: pd.Series, threshold: float) -> str:
    """[Doo 작업] 모델 원인을 단정하지 않고 관찰된 거래 상태를 요약합니다."""
    customer_type = customer["고객유형"]
    cycle_ratio = customer.get("평소_주기_대비")
    probability = customer["이탈확률"]

    if customer_type == "첫 구매 고객":
        return "첫 구매 이후 추가 구매가 확인되지 않아 재구매 유도 캠페인을 검토할 수 있습니다."
    if customer_type == "장기 구매 주기":
        return "구매 주기가 긴 고객이므로 일반 고객보다 더 긴 관찰 기간을 고려해야 합니다."
    if pd.notna(cycle_ratio) and cycle_ratio >= 1.5:
        return "평균 구매 주기를 크게 초과했고 이탈 확률도 높아 빠른 캠페인 검토가 필요합니다."
    if probability >= threshold:
        return "최종 운영 기준 이상으로 분류되었습니다. 거래 패턴과 캠페인 비용을 함께 검토해주세요."
    return "현재 캠페인 대상 기준 미만입니다. 향후 구매 활동 변화를 관찰해주세요."


def _build_campaign_export(
    source_df: pd.DataFrame,
    threshold: float,
    *,
    campaign_only: bool = True,
    data_as_of=None,
) -> pd.DataFrame:
    """[Doo 작업] 전체 캠페인 대상 또는 현재 필터 결과를 다운로드 표로 만듭니다."""
    export_df = source_df.copy()
    if campaign_only:
        export_df = export_df[export_df["이탈확률"] >= threshold].copy()
    export_df = export_df.sort_values("이탈확률", ascending=False)

    generated_at = datetime.now().astimezone()
    if data_as_of is None:
        data_as_of = source_df["last_purchase"].max()
    data_as_of_text = data_as_of.date().isoformat() if pd.notna(data_as_of) else ""

    export_df["추천캠페인"] = export_df["고객유형"].map(CAMPAIGN_BY_TYPE).fillna("일반 리텐션 캠페인")
    # [Doo 작업] XGBoost float32 확률을 변환해 CSV의 소수 오차를 제거합니다.
    export_df["이탈확률(%)"] = (export_df["이탈확률"].astype(float) * 100).round(1)
    export_df["적용기준(%)"] = round(threshold * 100, 1)
    export_df["취소·반품경험"] = export_df["has_return"].map({1: "있음", 0: "없음"})
    export_df["데이터기준일"] = data_as_of_text
    export_df["파일생성시각"] = generated_at.strftime("%Y-%m-%d %H:%M:%S %Z")

    export_columns = {
        "CustomerID": "CustomerID",
        "이탈확률(%)": "이탈확률(%)",
        "적용기준(%)": "적용기준(%)",
        "모델판정": "모델판정",
        "고객유형": "고객유형",
        "추천캠페인": "추천캠페인",
        "recency_days": "최근구매후경과일",
        "avg_days_between_orders": "평균구매간격(일)",
        "평소_주기_대비": "평소주기대비(배)",
        "frequency": "구매횟수",
        "distinct_products": "구매상품종류수",
        "net_revenue": "순매출",
        "취소·반품경험": "취소·반품경험",
        "데이터기준일": "데이터기준일",
        "파일생성시각": "파일생성시각",
    }
    result = export_df[list(export_columns)].rename(columns=export_columns).reset_index(drop=True)
    if not result.empty:
        result["CustomerID"] = result["CustomerID"].astype(int)
    result["평균구매간격(일)"] = result["평균구매간격(일)"].round(1)
    result["순매출"] = result["순매출"].round(2)
    return result


def _build_campaign_excel(campaign_df: pd.DataFrame, threshold: float, scope_label: str) -> bytes:
    """[Doo 작업] 담당자 검토용 서식과 사용 안내를 포함한 Excel 파일을 만듭니다."""
    buffer = BytesIO()
    guide_df = pd.DataFrame(
        {
            "항목": ["파일 목적", "다운로드 범위", "선정 기준", "고객 수", "데이터 출처", "사용 안내"],
            "내용": [
                "마케팅·CRM 고객 목록 검토",
                scope_label,
                f"현재 적용된 캠페인 기준 {threshold:.0%}",
                f"{len(campaign_df):,}명",
                "online_retail_II 거래 데이터를 고객 단위로 집계한 스냅샷",
                "CustomerID를 사내 CRM 고객 정보와 연결한 뒤 캠페인 발송에 사용하세요.",
            ],
        }
    )

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        campaign_df.to_excel(writer, sheet_name="고객 목록", index=False)
        guide_df.to_excel(writer, sheet_name="사용 안내", index=False)

        target_sheet = writer.sheets["고객 목록"]
        guide_sheet = writer.sheets["사용 안내"]
        header_fill = PatternFill("solid", fgColor="2F80ED")
        header_font = Font(color="FFFFFF", bold=True)

        # [Doo 작업] Excel 검토 편의를 위해 서식·필터·첫 행 고정을 적용합니다.
        for sheet in (target_sheet, guide_sheet):
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions

        for sheet in (target_sheet, guide_sheet):
            for column_index, column_cells in enumerate(sheet.columns, start=1):
                max_length = max(len(str(cell.value or "")) for cell in column_cells)
                sheet.column_dimensions[get_column_letter(column_index)].width = min(max_length + 3, 45)

        if not campaign_df.empty:
            probability_column = campaign_df.columns.get_loc("이탈확률(%)") + 1
            threshold_column = campaign_df.columns.get_loc("적용기준(%)") + 1
            revenue_column = campaign_df.columns.get_loc("순매출") + 1
            for row_index in range(2, len(campaign_df) + 2):
                target_sheet.cell(row=row_index, column=probability_column).number_format = '0.0"%"'
                target_sheet.cell(row=row_index, column=threshold_column).number_format = '0.0"%"'
                target_sheet.cell(row=row_index, column=revenue_column).number_format = '#,##0.00'

    buffer.seek(0)
    return buffer.getvalue()


def _kpi_card(column, label: str, value: str, description: str):
    """[Doo 작업] CRM 핵심 수치를 동일한 카드 형식으로 표시합니다."""
    with column:
        with st.container(border=True):
            st.caption(label)
            st.markdown(f"<div style='font-size:28px;font-weight:750'>{value}</div>", unsafe_allow_html=True)
            st.caption(description)


def _detail_row(label, value):
    value = "정보 없음" if pd.isna(value) else value
    st.markdown(
        f"""
        <div style="display:flex; justify-content:space-between; align-items:center;
                    padding:7px 0; border-bottom:1px solid #F0F0F0; gap:12px;">
            <span style="color:#667085; font-size:13px;">{label}</span>
            <span style="font-weight:700; font-size:14px; text-align:right;">{value}</span>
        </div>
        """,
        unsafe_allow_html=True,
    )


def _priority_badge(priority: str) -> str:
    background, color, icon = PRIORITY_COLORS[priority]
    return (
        f'<span style="background:{background};color:{color};font-weight:700;'
        f'padding:4px 10px;border-radius:14px;font-size:12px;">{icon} {priority}</span>'
    )


def _render_customer_detail(customer: pd.Series | None, threshold: float, data_as_of_text: str):
    """[Doo 작업] 선택 고객의 상태·위험 요인·추천 실행을 하나의 상세 카드로 연결합니다."""
    st.markdown("#### 선택 고객 상세 정보")
    if customer is None:
        st.info("고객 목록에서 고객을 선택하거나 CustomerID를 검색하면 상세 정보를 확인할 수 있습니다.")
        return

    priority = customer["우선순위"]
    campaign = _campaign_plan(customer)
    probability = customer["이탈확률"]
    customer_id = int(customer["CustomerID"])
    type_badge = BADGE_BY_TYPE.get(customer["고객유형"], "모니터링")

    with st.container(border=True):
        st.markdown(
            f"""
            <div style="display:flex;justify-content:space-between;align-items:flex-start;gap:12px;">
                <div>
                    <div style="color:#667085;font-size:12px;">CustomerID</div>
                    <div style="font-size:24px;font-weight:750;">{customer_id}</div>
                </div>
                <div style="text-align:right;">
                    <div style="font-size:30px;font-weight:750;">{probability:.1%}</div>
                    <div style="color:#667085;font-size:12px;">이탈 확률</div>
                </div>
            </div>
            <div style="margin:12px 0;display:flex;gap:8px;flex-wrap:wrap;">
                {_priority_badge(priority)}
                <span style="background:#F2F4F7;color:#344054;font-weight:650;
                             padding:4px 10px;border-radius:14px;font-size:12px;">
                    {customer['고객유형']}
                </span>
                <span style="background:#EEF4FF;color:#3538CD;font-weight:650;
                             padding:4px 10px;border-radius:14px;font-size:12px;">
                    {type_badge}
                </span>
            </div>
            """,
            unsafe_allow_html=True,
        )

        detail_left, detail_right = st.columns(2)
        with detail_left:
            _detail_row("최근 구매 후", f"{int(customer['recency_days'])}일")
            _detail_row("평균 구매 간격", f"{customer['avg_days_between_orders']:.1f}일")
            cycle_ratio = customer["평소_주기_대비"]
            _detail_row("평소 주기 대비", f"{cycle_ratio:.1f}배" if pd.notna(cycle_ratio) else "정보 없음")
            _detail_row("총 구매 횟수", f"{int(customer['frequency'])}회")
        with detail_right:
            _detail_row("구매 상품 종류", f"{int(customer['distinct_products']):,}개")
            _detail_row("순매출", f"£{customer['net_revenue']:,.2f}")
            _detail_row("취소·반품 경험", "있음" if customer["has_return"] == 1 else "없음")
            _detail_row("데이터 기준일", data_as_of_text or "정보 없음")

        st.info(_customer_interpretation(customer, threshold))

        st.markdown("**주요 위험 요인**")
        factors = _risk_factors(customer)
        if factors:
            st.markdown("\n".join(f"- {factor}" for factor in factors))
        else:
            st.caption(
                "거래 데이터에서 뚜렷한 단일 위험 요인은 확인되지 않았습니다. "
                "현재 이탈 확률과 전체 구매 패턴을 함께 검토해주세요."
            )

        st.markdown("**추천 캠페인**")
        st.markdown(f"**{campaign['name']}**")
        st.markdown(f"**추천 대상**  \n{campaign['target']}")
        st.caption(f"추천 이유 · {campaign['reason']}")
        st.markdown("**권장 실행 방법**")
        st.markdown("\n".join(f"- {action}" for action in campaign["actions"]))
        st.markdown(f"**기대 효과**  \n{campaign['effect']}")
        st.markdown(f"**운영 시 주의사항**  \n{campaign['caution']}")


def _clear_customer_search():
    """[Doo 작업] 검색어와 선택 고객을 함께 초기화합니다."""
    st.session_state["risk_search_input"] = ""
    st.session_state["selected_customer_id"] = None


def _render_download_group(
    title: str,
    description: str,
    export_df: pd.DataFrame,
    threshold: float,
    file_prefix: str,
    scope_label: str,
    key_prefix: str,
    *,
    csv_button_label: str = "CRM 발송 대상 CSV 다운로드",
):
    """[Doo 작업] 다운로드 범위와 파일 용도를 분리해 혼동을 방지합니다."""
    st.markdown(f"**{title}**")
    st.caption(description)

    if export_df.empty:
        st.info("현재 조건에 해당하는 고객이 없습니다. 필터 조건이나 캠페인 기준을 조정해주세요.")
        disabled_left, disabled_right = st.columns(2)
        disabled_left.download_button(
            "담당자 검토용 Excel 다운로드", data=b"", file_name="empty.xlsx",
            disabled=True, width="stretch", key=f"{key_prefix}_excel_empty",
        )
        disabled_right.download_button(
            csv_button_label, data=b"", file_name="empty.csv",
            disabled=True, width="stretch", key=f"{key_prefix}_csv_empty",
        )
        return

    file_date = datetime.now().astimezone().strftime("%Y%m%d")
    threshold_label = int(round(threshold * 100))
    try:
        excel_bytes = _build_campaign_excel(export_df, threshold, scope_label)
        csv_bytes = export_df.to_csv(index=False).encode("utf-8-sig")
    except Exception as error:
        st.error("파일을 생성하지 못했습니다. 데이터 상태를 확인한 뒤 다시 시도해주세요.")
        with st.expander("개발자 확인용 오류 정보"):
            st.code(str(error))
        return

    excel_column, csv_column = st.columns(2)
    excel_column.download_button(
        "담당자 검토용 Excel 다운로드",
        data=excel_bytes,
        file_name=f"{file_prefix}_threshold_{threshold_label}pct_{file_date}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        width="stretch",
        key=f"{key_prefix}_excel",
    )
    csv_column.download_button(
        csv_button_label,
        data=csv_bytes,
        file_name=f"{file_prefix}_threshold_{threshold_label}pct_{file_date}.csv",
        mime="text/csv",
        width="stretch",
        key=f"{key_prefix}_csv",
    )


def _prepare_scored_snapshot(model, preprocessor):
    """공통 고객 점수와 CRM 파생 컬럼을 한 번에 준비합니다."""
    snap = score_customers(load_customer_table(), model, preprocessor)
    threshold = DEFAULT_THRESHOLD
    snap["고객유형"] = snap.apply(segment, axis=1)
    snap["모델판정"] = snap["이탈확률"].ge(threshold).map(
        {True: "캠페인 대상 고객", False: "일반 관찰 고객"}
    )
    snap["평소_주기_대비"] = (
        snap["recency_days"] / snap["avg_days_between_orders"]
    ).replace([float("inf"), float("-inf")], pd.NA).round(2)
    snap["우선순위"] = snap["이탈확률"].apply(lambda value: _priority_label(value, threshold))
    snap["고객유형필터"] = snap.apply(_customer_type_group, axis=1)
    snap["추천캠페인"] = snap["고객유형"].map(CAMPAIGN_BY_TYPE).fillna("일반 리텐션 캠페인")
    return snap, threshold


def render_summary(model, preprocessor):
    """최종 운영 기준 38%의 대상 규모와 고객 구성을 표시합니다."""
    snap, threshold = _prepare_scored_snapshot(model, preprocessor)
    campaign_target = snap[snap["이탈확률"] >= threshold]

    st.divider()
    st.markdown(f"### 현재 고객 적용 결과 — 전체 고객 {len(snap):,}명 기준")
    st.caption("최종 운영 기준 38%를 현재 고객 스냅샷에 적용한 결과입니다.")

    first_kpi_row = st.columns(2)
    _kpi_card(first_kpi_row[0], "전체 고객", f"{len(snap):,}명", "현재 고객 스냅샷 전체")
    _kpi_card(
        first_kpi_row[1], "캠페인 대상 고객", f"{len(campaign_target):,}명",
        f"최종 운영 기준 {threshold:.0%} 이상",
    )
    second_kpi_row = st.columns(3)
    first_purchase_count = int((snap["고객유형"] == "첫 구매 고객").sum())
    long_cycle_count = int((snap["고객유형"] == "장기 구매 주기").sum())
    average_target_probability = campaign_target["이탈확률"].mean() if not campaign_target.empty else 0
    _kpi_card(second_kpi_row[0], "첫 구매 고객", f"{first_purchase_count:,}명", "구매 횟수 1회")
    _kpi_card(second_kpi_row[1], "장기 구매 주기 고객", f"{long_cycle_count:,}명", "평균 구매 간격 90일 이상")
    _kpi_card(
        second_kpi_row[2], "대상 고객 평균 이탈 확률", f"{average_target_probability:.1%}",
        "현재 캠페인 대상 고객 평균",
    )

    summary_left, summary_right = st.columns([1.15, 1], gap="large")
    with summary_left:
        st.markdown("#### 고객 선정 구성")
        selection_summary = pd.DataFrame(
            {
                "고객 수": [
                    len(campaign_target),
                    len(snap) - len(campaign_target),
                ],
            },
            index=["캠페인 대상", "관찰 고객"],
        )
        st.bar_chart(selection_summary, horizontal=True, color="#2F80ED")
    with summary_right:
        st.markdown("#### 다음 작업")
        st.info(
            "선정 규모가 적절하면 **고객 목록** 메뉴에서 대상 고객을 검색·필터링하고 "
            "Excel 또는 CSV로 내려받으세요."
        )
        st.markdown(
            f"- 최종 운영 기준: **{threshold:.0%}**\n"
            f"- 캠페인 대상 비율: **{len(campaign_target) / len(snap):.1%}**\n"
            f"- 관찰 고객 비율: **{(len(snap) - len(campaign_target)) / len(snap):.1%}**"
        )


def render(model, preprocessor):
    """검색·필터·테이블·상세·다운로드 중심의 고객 목록 화면을 렌더링합니다."""
    snap, threshold = _prepare_scored_snapshot(model, preprocessor)

    if "selected_customer_id" not in st.session_state:
        st.session_state["selected_customer_id"] = None
    if "risk_search_input" not in st.session_state:
        st.session_state["risk_search_input"] = ""

    data_as_of = snap["last_purchase"].max()
    data_as_of_text = data_as_of.date().isoformat() if pd.notna(data_as_of) else ""
    campaign_target = snap[snap["이탈확률"] >= threshold]

    st.markdown("### 고객 목록")
    st.markdown(
        "최종 운영 기준 38%를 바탕으로 고객을 검색·필터링합니다.  \n"
        "행을 선택하면 상세 정보와 추천 캠페인을 확인하고 실행용 목록을 내려받을 수 있습니다."
    )
    st.markdown(
        f"""
        <div style="background:#EAF3FF;border:1px solid #BBD7FF;border-radius:10px;
                    padding:14px 18px;margin:12px 0 18px 0;">
            <div style="color:#1F6FCC;font-size:13px;font-weight:700;">최종 운영 기준</div>
            <div style="font-size:24px;font-weight:750;margin:4px 0;">{threshold:.0%}</div>
            <div style="color:#425466;font-size:13px;">
                이탈 확률이 {threshold:.0%} 이상인 고객이 캠페인 대상입니다.
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    _render_campaign_guide()

    # CustomerID 검색 — 필터보다 먼저 제공하고 검색 결과는 필터와 무관하게 선택합니다.
    st.markdown("#### CustomerID 검색")
    search_column, clear_column = st.columns([5, 1])
    with search_column:
        with st.form("risk_customer_search_form"):
            search_query = st.text_input(
                "CustomerID",
                placeholder="예: 17850",
                key="risk_search_input",
                label_visibility="collapsed",
            )
            search_submitted = st.form_submit_button("고객 검색", type="primary", width="stretch")
    with clear_column:
        st.button("검색 초기화", on_click=_clear_customer_search, width="stretch")

    if search_submitted:
        try:
            searched_customer_id = int(search_query.strip())
            matched = snap[snap["CustomerID"] == searched_customer_id]
        except (ValueError, AttributeError):
            matched = pd.DataFrame()
        if matched.empty:
            st.error("해당 CustomerID를 찾을 수 없습니다. 입력한 번호를 다시 확인해주세요.")
        else:
            st.session_state["selected_customer_id"] = searched_customer_id
            st.success(f"CustomerID {searched_customer_id} 고객을 선택했습니다. 필터와 관계없이 상세 정보를 표시합니다.")

    # 3) 관리 범위와 고객 유형을 독립적으로 적용합니다.
    st.markdown("#### 고객 필터")
    management_counts = {
        "전체 고객": len(snap),
        "캠페인 대상 고객": len(campaign_target),
    }
    type_options = ["전체 유형", "첫 구매 고객", "반복 구매 고객", "장기 구매 주기 고객", "기타 관리 고객"]

    management_column, type_column = st.columns(2)
    management_filter = management_column.selectbox(
        "관리 범위",
        list(management_counts),
        format_func=lambda option: f"{option}  {management_counts[option]:,}명",
        key="risk_management_filter",
    )
    type_filter = type_column.selectbox(
        "고객 유형",
        type_options,
        # [Doo 작업] 고객 유형 드롭다운은 인원 수 없이 유형명만 간결하게 표시합니다.
        key="risk_type_filter",
    )

    view = snap
    if management_filter == "캠페인 대상 고객":
        view = view[view["이탈확률"] >= threshold]
    if type_filter != "전체 유형":
        view = view[view["고객유형필터"] == type_filter]
    view = view.sort_values("이탈확률", ascending=False).reset_index(drop=True)

    # 4) 목록과 상세 정보를 65:35로 연결합니다.
    list_column, detail_column = st.columns([1.65, 1], gap="large")
    with list_column:
        st.markdown(f"#### 우선순위 고객 목록 · {len(view):,}명")
        st.caption("이탈 확률이 높은 순으로 표시됩니다. 표의 열 제목을 눌러 다시 정렬할 수 있습니다.")
        if view.empty:
            st.info("현재 조건에 해당하는 고객이 없습니다. 관리 범위나 고객 유형 필터를 확인해주세요.")
        else:
            display_df = view[
                ["CustomerID", "고객유형", "이탈확률", "평소_주기_대비", "추천캠페인", "우선순위"]
            ].copy()
            display_df.insert(0, "등급", display_df["우선순위"].map(
                lambda priority: f"{PRIORITY_COLORS[priority][2]} {priority}"
            ))
            display_df = display_df.drop(columns="우선순위")
            display_df["이탈확률"] = (display_df["이탈확률"].astype(float) * 100).round(1)
            display_df["평소_주기_대비"] = display_df["평소_주기_대비"].astype(float).round(1)

            table_event = st.dataframe(
                display_df,
                hide_index=True,
                width="stretch",
                height=500,
                on_select="rerun",
                selection_mode="single-row",
                key="risk_customer_table",
                column_config={
                    "등급": st.column_config.TextColumn("우선순위", width="small"),
                    "CustomerID": st.column_config.NumberColumn("CustomerID", format="%d"),
                    "고객유형": st.column_config.TextColumn("고객 유형", width="medium"),
                    "이탈확률": st.column_config.ProgressColumn(
                        "이탈 확률", format="%.1f%%", min_value=0, max_value=100,
                    ),
                    "평소_주기_대비": st.column_config.NumberColumn("평소 주기 대비", format="%.1f배"),
                    "추천캠페인": st.column_config.TextColumn("추천 캠페인", width="medium"),
                },
            )
            selected_rows = table_event.selection.rows if table_event.selection else []
            if selected_rows:
                selected_customer_id = int(view.iloc[selected_rows[0]]["CustomerID"])
                st.session_state["selected_customer_id"] = selected_customer_id

    selected_customer = None
    selected_customer_id = st.session_state.get("selected_customer_id")
    if selected_customer_id is not None:
        matched_customer = snap[snap["CustomerID"] == int(selected_customer_id)]
        if not matched_customer.empty:
            selected_customer = matched_customer.iloc[0]
    with detail_column:
        _render_customer_detail(selected_customer, threshold, data_as_of_text)

    # 5) 다운로드 범위를 현재 필터 결과와 전체 캠페인 대상으로 명확히 분리합니다.
    st.divider()
    st.markdown("### 고객 목록 다운로드")
    st.markdown(
        f"최종 운영 기준 **{threshold:.0%} 이상**, 전체 캠페인 대상 고객 **{len(campaign_target):,}명**입니다.  \n"
        "조회·검토용 목록에는 일반 고객이 포함될 수 있습니다. 실제 발송에는 캠페인 실행 대상 파일만 사용하세요."
    )

    filtered_export = _build_campaign_export(
        view, threshold, campaign_only=False, data_as_of=data_as_of,
    )
    full_campaign_export = _build_campaign_export(
        snap, threshold, campaign_only=True, data_as_of=data_as_of,
    )
    download_summary = st.columns(2)
    download_summary[0].metric("현재 조회 고객", f"{len(filtered_export):,}명")
    download_summary[1].metric("캠페인 실행 대상", f"{len(full_campaign_export):,}명")

    filtered_download_column, full_download_column = st.columns(2, gap="large")
    with filtered_download_column:
        with st.container(border=True):
            _render_download_group(
                "① 조회·검토용 목록",
                f"현재 화면 필터를 만족하는 {len(filtered_export):,}명입니다. 38% 미만 일반 고객이 포함될 수 있으며 실제 발송용이 아닙니다.",
                filtered_export,
                threshold,
                "filtered_customers",
                "현재 화면 필터 결과",
                "filtered_customers",
                csv_button_label="조회 결과 CSV 다운로드",
            )
    with full_download_column:
        with st.container(border=True):
            _render_download_group(
                "② 캠페인 실행 대상",
                f"화면 필터와 관계없이 최종 운영 기준 38% 이상 고객 {len(full_campaign_export):,}명만 포함합니다. 실제 CRM 발송에는 이 파일을 사용하세요.",
                full_campaign_export,
                threshold,
                "campaign_targets",
                "전체 캠페인 대상",
                "campaign_targets",
            )
